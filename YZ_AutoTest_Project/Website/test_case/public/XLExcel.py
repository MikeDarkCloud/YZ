from openpyxl import *
import os
class rewrxl():

    def write_xl(self,col,value):
        pwd = os.getcwd()
        p1 = os.path.abspath(os.path.dirname(pwd) + os.path.sep + "..")
        p2 = "test_data\Excel_File\StudentBaseInfo.xlsx"
        file = os.path.join(pwd, p2)
        wb = load_workbook(file)
        #打开sheet
        ws = wb["StudentBaseInfo"]
        #方法1
        # ws.cell(row=3, column=3).value = 'test'
        #方法2
        row=ws.max_row
        if col=='A':
            ws['%s%d' %(col,row+1)]= value
        else:
            ws['%s%d' % (col, row)] = value
        wb.save(file)

    def read_xl(self,e_loc):
        pwd = os.getcwd()
        p1 = os.path.abspath(os.path.dirname(pwd) + os.path.sep + "..")
        p2 = "test_data\Excel_File\StudentBaseInfo.xlsx"
        file = os.path.join(pwd, p2)
        wb = load_workbook(file)
        ws = wb.get_sheet_by_name("Login_Eelement")
        ss=ws[e_loc].value
        return ss

    def sread_xl(self):
        pwd = os.getcwd()
        # p1 = os.path.abspath(os.path.dirname(pwd) + os.path.sep + "..")
        p2 = "test_data\Excel_File\StudentBaseInfo.xlsx"
        file = os.path.join(pwd, p2)
        wb = load_workbook(file)
        ws = wb.get_sheet_by_name("StudentNum")
        ss=ws['A1'].value
        return ss

    def swrite_xl(self,nvalue):
        pwd = os.getcwd()
        p1 = os.path.abspath(os.path.dirname(pwd) + os.path.sep + "..")
        p2 = "test_data\Excel_File\StudentBaseInfo.xlsx"
        file = os.path.join(pwd, p2)
        wb = load_workbook(file)
        ws = wb["StudentNum"]
        ws['%s%d' % ('A', 1)] = nvalue
        wb.save(file)

if __name__ == '__main__':
    l=rewrxl()
    l.write_xl('A','123456')
    # print(l.sread_xl())
    # l.swrite_xl(2)
    # write_xl()
    # read_xl()
